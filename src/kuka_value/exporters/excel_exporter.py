"""Excel (.xlsx) export for robot analysis results."""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font

from kuka_value.exporters.base import Exporter
from kuka_value.models.axis_load import AxisLoad
from kuka_value.models.payload import Payload
from kuka_value.models.robot_info import RobotInfo

PAYLOAD_HEADERS = [
    "Index(es)",
    "Mass (kg)",
    "CoG X (mm)",
    "CoG Y (mm)",
    "CoG Z (mm)",
    "Orientation A (deg)",
    "Orientation B (deg)",
    "Orientation C (deg)",
    "Inertia X (kgm2)",
    "Inertia Y (kgm2)",
    "Inertia Z (kgm2)",
    "Source File",
]

AXIS_LOAD_HEADERS = [
    "Axis",
    "Mass (kg)",
    "CoG X (mm)",
    "CoG Y (mm)",
    "CoG Z (mm)",
    "Orientation A (deg)",
    "Orientation B (deg)",
    "Orientation C (deg)",
    "Inertia X (kgm2)",
    "Inertia Y (kgm2)",
    "Inertia Z (kgm2)",
    "Source File",
]

_BOLD = Font(bold=True)


class ExcelExporter(Exporter):
    """Exports a robot analysis result as a formatted .xlsx workbook.

    Produces three sheets: "Summary" (model/backup/controller
    metadata), "Payloads" (the unique-payload table), and "Axis Loads"
    (supplementary per-axis loads, e.g. LOAD_A3_DATA).
    """

    def export(self, robot: RobotInfo) -> bytes:
        workbook = Workbook()
        self._write_summary_sheet(workbook, robot)
        self._write_payloads_sheet(workbook, robot)
        self._write_axis_loads_sheet(workbook, robot)

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def _write_summary_sheet(workbook: Workbook, robot: RobotInfo) -> None:
        sheet = workbook.active
        assert sheet is not None
        sheet.title = "Summary"

        rows: list[tuple[str, str | int]] = [
            ("Model", robot.model),
            ("Backup Name", robot.general.backup_name),
            ("KSS Version", robot.general.kss_version or ""),
            ("Controller Type", robot.controller.controller_type.value),
            ("Serial Number", robot.controller.serial_number or ""),
            ("Unique Payloads", len(robot.payloads)),
            ("Axis Loads", len(robot.axis_loads)),
            ("Warnings", len(robot.warnings)),
        ]
        for row in rows:
            sheet.append(row)

        for cell in sheet["A"]:
            cell.font = _BOLD

    @staticmethod
    def _write_payloads_sheet(workbook: Workbook, robot: RobotInfo) -> None:
        sheet = workbook.create_sheet("Payloads")
        sheet.append(PAYLOAD_HEADERS)
        for cell in sheet[1]:
            cell.font = _BOLD

        for payload in robot.payloads:
            sheet.append(ExcelExporter.payload_row(payload))

    @staticmethod
    def _write_axis_loads_sheet(workbook: Workbook, robot: RobotInfo) -> None:
        sheet = workbook.create_sheet("Axis Loads")
        sheet.append(AXIS_LOAD_HEADERS)
        for cell in sheet[1]:
            cell.font = _BOLD

        for axis_load in robot.axis_loads:
            sheet.append(ExcelExporter.axis_load_row(axis_load))

    @staticmethod
    def payload_row(payload: Payload) -> list[str | float | None]:
        inertia = payload.inertia
        orientation = payload.orientation
        return [
            ", ".join(str(i) for i in payload.indices),
            payload.mass,
            payload.center_of_gravity.x,
            payload.center_of_gravity.y,
            payload.center_of_gravity.z,
            orientation.a if orientation else None,
            orientation.b if orientation else None,
            orientation.c if orientation else None,
            inertia.x if inertia else None,
            inertia.y if inertia else None,
            inertia.z if inertia else None,
            payload.source_file or "",
        ]

    @staticmethod
    def axis_load_row(axis_load: AxisLoad) -> list[str | float | int | None]:
        inertia = axis_load.inertia
        orientation = axis_load.orientation
        return [
            axis_load.axis,
            axis_load.mass,
            axis_load.center_of_gravity.x,
            axis_load.center_of_gravity.y,
            axis_load.center_of_gravity.z,
            orientation.a if orientation else None,
            orientation.b if orientation else None,
            orientation.c if orientation else None,
            inertia.x if inertia else None,
            inertia.y if inertia else None,
            inertia.z if inertia else None,
            axis_load.source_file or "",
        ]
