"""Excel (.xlsx) export for a batch of robot analysis results."""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font

from kuka_value.exporters.batch_base import BatchExporter
from kuka_value.exporters.excel_exporter import (
    AXIS_LOAD_HEADERS,
    PAYLOAD_HEADERS,
    ExcelExporter,
)
from kuka_value.models.batch_result import BatchItemResult

_SUMMARY_HEADERS = ["Backup Name", "Model", "Payloads", "Axis Loads", "Warnings", "Status"]
_BOLD = Font(bold=True)


class BatchExcelExporter(BatchExporter):
    """Exports a batch of results as a formatted .xlsx workbook.

    Produces three sheets: "Summary" (one row per backup, including
    failures), "Payloads", and "Axis Loads" - the latter two combined
    tables across all successfully analyzed backups, prefixed with
    Backup Name/Model.
    """

    def export(self, results: list[BatchItemResult]) -> bytes:
        workbook = Workbook()
        self._write_summary_sheet(workbook, results)
        self._write_payloads_sheet(workbook, results)
        self._write_axis_loads_sheet(workbook, results)

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def _write_summary_sheet(workbook: Workbook, results: list[BatchItemResult]) -> None:
        sheet = workbook.active
        assert sheet is not None
        sheet.title = "Summary"

        sheet.append(_SUMMARY_HEADERS)
        for cell in sheet[1]:
            cell.font = _BOLD

        for result in results:
            sheet.append(BatchExcelExporter._summary_row(result))

    @staticmethod
    def _summary_row(result: BatchItemResult) -> list[str | int]:
        if result.robot is None:
            return [result.display_name, "-", 0, 0, 0, f"FAILED: {result.error}"]
        return [
            result.display_name,
            result.robot.model,
            len(result.robot.payloads),
            len(result.robot.axis_loads),
            len(result.robot.warnings),
            "OK",
        ]

    @staticmethod
    def _write_payloads_sheet(workbook: Workbook, results: list[BatchItemResult]) -> None:
        sheet = workbook.create_sheet("Payloads")
        sheet.append(["Backup Name", "Model", *PAYLOAD_HEADERS])
        for cell in sheet[1]:
            cell.font = _BOLD

        for result in results:
            if result.robot is None:
                continue
            for payload in result.robot.payloads:
                sheet.append(
                    [result.display_name, result.robot.model, *ExcelExporter.payload_row(payload)]
                )

    @staticmethod
    def _write_axis_loads_sheet(workbook: Workbook, results: list[BatchItemResult]) -> None:
        sheet = workbook.create_sheet("Axis Loads")
        sheet.append(["Backup Name", "Model", *AXIS_LOAD_HEADERS])
        for cell in sheet[1]:
            cell.font = _BOLD

        for result in results:
            if result.robot is None:
                continue
            for axis_load in result.robot.axis_loads:
                sheet.append(
                    [
                        result.display_name,
                        result.robot.model,
                        *ExcelExporter.axis_load_row(axis_load),
                    ]
                )
