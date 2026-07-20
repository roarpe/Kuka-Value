"""Integration tests for BatchResultsWindow.

Instantiates the real widget tree via pytest-qt's qtbot fixture and
drives real backups through Engine.parse_many() end-to-end (a genuine
QThread, not a synchronous shortcut), same approach as
test_main_window.py.

BatchResultsWindow starts its background QThread from __init__, and
worker.all_finished -> thread.quit() is a cross-thread queued
connection: it only actually runs once the main thread's Qt event
loop gets pumped (which qtbot.waitUntil does internally). Every test
that constructs a window must wait for the thread to finish, or the
QThread is left running when the test ends and pytest hangs at
teardown waiting for it.
"""

import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from kuka_value.engine.engine import Engine
from kuka_value.exporters.batch_csv_exporter import BatchCsvExporter
from kuka_value.exporters.batch_json_exporter import BatchJsonExporter
from kuka_value.ui.batch_results_window import BatchResultsWindow


@pytest.fixture
def temp_dir() -> Path:
    with tempfile.TemporaryDirectory() as tmpdir:
        yield Path(tmpdir)


@pytest.fixture
def engine() -> Engine:
    return Engine()


def _make_backup(folder: Path, name: str, trafoname: str) -> Path:
    backup = folder / name
    backup.mkdir()
    (backup / "$machine.dat").write_text(f'$TRAFONAME[]="{trafoname}"\n')
    return backup


def _wait_for_batch_done(window: BatchResultsWindow, qtbot: object) -> None:
    qtbot.waitUntil(lambda: window._thread is None, timeout=5000)  # type: ignore[attr-defined]


class TestBatchExecutionEndToEnd:
    def test_analyzes_all_backups_and_populates_summary(
        self, engine: Engine, temp_dir: Path, qtbot: object
    ) -> None:
        backups = [
            _make_backup(temp_dir, "Robot1", "KR240R2900"),
            _make_backup(temp_dir, "Robot2", "KR6R900"),
        ]

        window = BatchResultsWindow(engine, backups)
        qtbot.addWidget(window)  # type: ignore[attr-defined]
        _wait_for_batch_done(window, qtbot)

        assert window._summary_model.rowCount() == 2
        assert window._summary_model.result_at(0).robot is not None
        assert window._summary_model.result_at(0).robot.model == "KR 240 R2900"
        assert window._summary_model.result_at(1).robot.model == "KR 6 R900"

    def test_isolates_one_bad_backup_among_good_ones(
        self, engine: Engine, temp_dir: Path, qtbot: object
    ) -> None:
        good = _make_backup(temp_dir, "GoodRobot", "KR240R2900")
        missing = temp_dir / "does_not_exist"

        window = BatchResultsWindow(engine, [good, missing])
        qtbot.addWidget(window)  # type: ignore[attr-defined]
        _wait_for_batch_done(window, qtbot)

        assert window._summary_model.result_at(0).succeeded is True
        assert window._summary_model.result_at(1).succeeded is False

    def test_export_actions_enabled_after_completion(
        self, engine: Engine, temp_dir: Path, qtbot: object
    ) -> None:
        backup = _make_backup(temp_dir, "Robot1", "KR240R2900")

        window = BatchResultsWindow(engine, [backup])
        qtbot.addWidget(window)  # type: ignore[attr-defined]

        assert window.action_export_csv.isEnabled() is False

        _wait_for_batch_done(window, qtbot)
        assert window.action_export_csv.isEnabled()
        assert window.action_export_excel.isEnabled()
        assert window.action_export_json.isEnabled()

    def test_status_bar_reports_completion_counts(
        self, engine: Engine, temp_dir: Path, qtbot: object
    ) -> None:
        good = _make_backup(temp_dir, "GoodRobot", "KR240R2900")
        missing = temp_dir / "does_not_exist"

        window = BatchResultsWindow(engine, [good, missing])
        qtbot.addWidget(window)  # type: ignore[attr-defined]
        _wait_for_batch_done(window, qtbot)

        assert "1 succeeded" in window.statusBar().currentMessage()
        assert "1 failed" in window.statusBar().currentMessage()

    def test_empty_batch_completes_without_hanging(self, engine: Engine, qtbot: object) -> None:
        window = BatchResultsWindow(engine, [])
        qtbot.addWidget(window)  # type: ignore[attr-defined]
        _wait_for_batch_done(window, qtbot)

        assert window._summary_model.rowCount() == 0


class TestSelection:
    def test_clicking_summary_row_populates_payload_table(
        self, engine: Engine, temp_dir: Path, qtbot: object
    ) -> None:
        backup = temp_dir / "Robot1"
        backup.mkdir()
        (backup / "$machine.dat").write_text('$TRAFONAME[]="KR240R2900"\n')
        (backup / "$config.dat").write_text(
            "$LOAD_DATA[1]={M 10.5,CM {X 1.0,Y 2.0,Z 3.0,A 0.0,B 0.0,C 0.0},"
            "J {X 0.1,Y 0.1,Z 0.1}}\n"
        )

        window = BatchResultsWindow(engine, [backup])
        qtbot.addWidget(window)  # type: ignore[attr-defined]
        _wait_for_batch_done(window, qtbot)

        window._on_summary_row_clicked(window._summary_model.index(0, 0))

        assert window._payload_model.rowCount() == 1

    def test_clicking_failed_row_shows_empty_payload_table(
        self, engine: Engine, temp_dir: Path, qtbot: object
    ) -> None:
        missing = temp_dir / "does_not_exist"

        window = BatchResultsWindow(engine, [missing])
        qtbot.addWidget(window)  # type: ignore[attr-defined]
        _wait_for_batch_done(window, qtbot)

        window._on_summary_row_clicked(window._summary_model.index(0, 0))

        assert window._payload_model.rowCount() == 0


class TestExport:
    def test_export_with_no_results_does_nothing(
        self, engine: Engine, qtbot: object, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        window = BatchResultsWindow(engine, [])
        qtbot.addWidget(window)  # type: ignore[attr-defined]
        _wait_for_batch_done(window, qtbot)

        dialog_calls = []
        monkeypatch.setattr(
            "kuka_value.ui.batch_results_window.QFileDialog.getSaveFileName",
            lambda *_a: dialog_calls.append(True),
        )

        window._on_export_csv()

        assert dialog_calls == []

    def test_export_csv_writes_expected_content(
        self,
        engine: Engine,
        temp_dir: Path,
        tmp_path: Path,
        qtbot: object,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        backup = _make_backup(temp_dir, "Robot1", "KR240R2900")
        target = tmp_path / "batch.csv"

        window = BatchResultsWindow(engine, [backup])
        qtbot.addWidget(window)  # type: ignore[attr-defined]
        _wait_for_batch_done(window, qtbot)

        monkeypatch.setattr(
            "kuka_value.ui.batch_results_window.QFileDialog.getSaveFileName",
            lambda *_a: (str(target), "CSV Files (*.csv)"),
        )

        window._on_export_csv()

        assert target.exists()
        assert target.read_bytes() == BatchCsvExporter().export(window._results)

    def test_export_excel_writes_expected_content(
        self,
        engine: Engine,
        temp_dir: Path,
        tmp_path: Path,
        qtbot: object,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        backup = _make_backup(temp_dir, "Robot1", "KR240R2900")
        target = tmp_path / "batch.xlsx"

        window = BatchResultsWindow(engine, [backup])
        qtbot.addWidget(window)  # type: ignore[attr-defined]
        _wait_for_batch_done(window, qtbot)

        monkeypatch.setattr(
            "kuka_value.ui.batch_results_window.QFileDialog.getSaveFileName",
            lambda *_a: (str(target), "Excel Files (*.xlsx)"),
        )

        window._on_export_excel()

        assert target.exists()
        workbook = load_workbook(target)
        assert workbook.sheetnames == ["Summary", "Payloads", "Axis Loads"]

    def test_export_json_writes_expected_content(
        self,
        engine: Engine,
        temp_dir: Path,
        tmp_path: Path,
        qtbot: object,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        backup = _make_backup(temp_dir, "Robot1", "KR240R2900")
        target = tmp_path / "batch.json"

        window = BatchResultsWindow(engine, [backup])
        qtbot.addWidget(window)  # type: ignore[attr-defined]
        _wait_for_batch_done(window, qtbot)

        monkeypatch.setattr(
            "kuka_value.ui.batch_results_window.QFileDialog.getSaveFileName",
            lambda *_a: (str(target), "JSON Files (*.json)"),
        )

        window._on_export_json()

        assert target.exists()
        assert target.read_bytes() == BatchJsonExporter().export(window._results)

    def test_export_oserror_shows_error_dialog(
        self,
        engine: Engine,
        temp_dir: Path,
        tmp_path: Path,
        qtbot: object,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        backup = _make_backup(temp_dir, "Robot1", "KR240R2900")
        target = tmp_path / "batch.csv"

        window = BatchResultsWindow(engine, [backup])
        qtbot.addWidget(window)  # type: ignore[attr-defined]
        _wait_for_batch_done(window, qtbot)

        monkeypatch.setattr(
            "kuka_value.ui.batch_results_window.QFileDialog.getSaveFileName",
            lambda *_a: (str(target), "CSV Files (*.csv)"),
        )

        def boom(_self: object, _results: object, _path: object) -> None:
            raise OSError("disk full")

        monkeypatch.setattr(
            "kuka_value.exporters.batch_csv_exporter.BatchCsvExporter.export_to_file", boom
        )

        shown_messages = []
        monkeypatch.setattr(
            "kuka_value.ui.batch_results_window.QMessageBox.critical",
            lambda *args: shown_messages.append(args[-1]),
        )

        window._on_export_csv()

        assert shown_messages == ["disk full"]


class TestFatalError:
    def test_fatal_error_shows_message_box(
        self, engine: Engine, qtbot: object, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        window = BatchResultsWindow(engine, [])
        qtbot.addWidget(window)  # type: ignore[attr-defined]
        _wait_for_batch_done(window, qtbot)

        shown_messages = []
        monkeypatch.setattr(
            "kuka_value.ui.batch_results_window.QMessageBox.critical",
            lambda *args: shown_messages.append(args[-1]),
        )

        window._on_fatal_error("catastrophic failure")

        assert shown_messages == ["catastrophic failure"]


class TestBatchPathDiscovery:
    def test_discovers_zip_files_and_subfolders(self, temp_dir: Path) -> None:
        from kuka_value.ui.main_window import discover_batch_paths

        (temp_dir / "Robot1").mkdir()
        (temp_dir / "Robot2.zip").write_bytes(b"fake zip content")
        (temp_dir / "loose_file.txt").write_text("not a backup")

        paths = discover_batch_paths(temp_dir)

        assert sorted(p.name for p in paths) == ["Robot1", "Robot2.zip"]

    def test_empty_folder_yields_no_paths(self, temp_dir: Path) -> None:
        from kuka_value.ui.main_window import discover_batch_paths

        assert discover_batch_paths(temp_dir) == []
