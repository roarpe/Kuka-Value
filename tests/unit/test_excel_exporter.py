"""Unit tests for ExcelExporter."""

import io
from pathlib import Path

from openpyxl import load_workbook

from kuka_value.exporters.excel_exporter import ExcelExporter
from kuka_value.models.axis_load import AxisLoad
from kuka_value.models.controller_info import ControllerInfo, ControllerType
from kuka_value.models.general_info import GeneralInfo
from kuka_value.models.payload import Vector3D
from kuka_value.models.robot_info import RobotInfo
from kuka_value.models.warnings import WarningLog


class TestExcelExport:
    """Test Excel workbook structure and content."""

    def test_export_returns_bytes(self, sample_robot_info: RobotInfo) -> None:
        result = ExcelExporter().export(sample_robot_info)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_export_creates_three_sheets(self, sample_robot_info: RobotInfo) -> None:
        result = ExcelExporter().export(sample_robot_info)
        workbook = load_workbook(io.BytesIO(result))

        assert workbook.sheetnames == ["Summary", "Payloads", "Axis Loads"]

    def test_summary_sheet_contains_metadata(self, sample_robot_info: RobotInfo) -> None:
        result = ExcelExporter().export(sample_robot_info)
        workbook = load_workbook(io.BytesIO(result))
        summary = workbook["Summary"]

        values = {row[0]: row[1] for row in summary.iter_rows(values_only=True)}
        assert values["Model"] == "KR 240 R2900"
        assert values["Backup Name"] == "TestBackup"
        assert values["KSS Version"] == "8.6.8"
        assert values["Controller Type"] == "KRC4"
        assert values["Serial Number"] == "12345"
        assert values["Unique Payloads"] == 2
        assert values["Axis Loads"] == 0
        assert values["Warnings"] == 1

    def test_payloads_sheet_header_row(self, sample_robot_info: RobotInfo) -> None:
        result = ExcelExporter().export(sample_robot_info)
        workbook = load_workbook(io.BytesIO(result))
        sheet = workbook["Payloads"]

        header = [cell.value for cell in sheet[1]]
        assert header[0] == "Index(es)"
        assert header[1] == "Mass (kg)"

    def test_payloads_sheet_data_rows(self, sample_robot_info: RobotInfo) -> None:
        result = ExcelExporter().export(sample_robot_info)
        workbook = load_workbook(io.BytesIO(result))
        sheet = workbook["Payloads"]

        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 2

        first = rows[0]
        assert first[0] == "1, 3"
        assert first[1] == 10.5
        assert first[2] == 100.0
        assert first[5] == 0.5  # inertia X

    def test_payloads_sheet_missing_inertia_is_none(self, sample_robot_info: RobotInfo) -> None:
        result = ExcelExporter().export(sample_robot_info)
        workbook = load_workbook(io.BytesIO(result))
        sheet = workbook["Payloads"]

        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        second = rows[1]
        assert second[1] == 25.0
        assert second[5] is None
        assert second[6] is None
        assert second[7] is None

    def test_header_cells_are_bold(self, sample_robot_info: RobotInfo) -> None:
        result = ExcelExporter().export(sample_robot_info)
        workbook = load_workbook(io.BytesIO(result))
        sheet = workbook["Payloads"]

        assert sheet["A1"].font.bold is True

    def test_export_empty_payloads_produces_header_only(self) -> None:
        robot = RobotInfo(
            model="UNKNOWN",
            general=GeneralInfo(backup_name="Empty"),
            controller=ControllerInfo(controller_type=ControllerType.UNKNOWN),
            payloads=[],
            warnings=WarningLog(),
        )
        result = ExcelExporter().export(robot)
        workbook = load_workbook(io.BytesIO(result))
        sheet = workbook["Payloads"]

        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        assert rows == []


class TestAxisLoadsSheet:
    def test_no_axis_loads_produces_header_only(self, sample_robot_info: RobotInfo) -> None:
        result = ExcelExporter().export(sample_robot_info)
        workbook = load_workbook(io.BytesIO(result))
        sheet = workbook["Axis Loads"]

        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        assert rows == []

    def test_header_row(self) -> None:
        robot = RobotInfo(
            model="KR 240 R2900",
            general=GeneralInfo(backup_name="Test"),
            controller=ControllerInfo(controller_type=ControllerType.UNKNOWN),
            axis_loads=[
                AxisLoad(axis=3, mass=12.5, center_of_gravity=Vector3D(x=0.0, y=0.0, z=0.0))
            ],
            warnings=WarningLog(),
        )
        result = ExcelExporter().export(robot)
        workbook = load_workbook(io.BytesIO(result))
        sheet = workbook["Axis Loads"]

        header = [cell.value for cell in sheet[1]]
        assert header[0] == "Axis"
        assert header[1] == "Mass (kg)"

    def test_data_row_values(self) -> None:
        robot = RobotInfo(
            model="KR 240 R2900",
            general=GeneralInfo(backup_name="Test"),
            controller=ControllerInfo(controller_type=ControllerType.UNKNOWN),
            axis_loads=[
                AxisLoad(
                    axis=3,
                    mass=12.5,
                    center_of_gravity=Vector3D(x=50.0, y=0.0, z=0.0),
                    inertia=Vector3D(x=0.1, y=0.2, z=0.3),
                    source_file="$config.dat",
                )
            ],
            warnings=WarningLog(),
        )
        result = ExcelExporter().export(robot)
        workbook = load_workbook(io.BytesIO(result))
        sheet = workbook["Axis Loads"]

        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        assert rows[0] == (3, 12.5, 50.0, 0.0, 0.0, 0.1, 0.2, 0.3, "$config.dat")

    def test_missing_inertia_is_none(self) -> None:
        robot = RobotInfo(
            model="KR 240 R2900",
            general=GeneralInfo(backup_name="Test"),
            controller=ControllerInfo(controller_type=ControllerType.UNKNOWN),
            axis_loads=[
                AxisLoad(axis=1, mass=5.0, center_of_gravity=Vector3D(x=0.0, y=0.0, z=0.0))
            ],
            warnings=WarningLog(),
        )
        result = ExcelExporter().export(robot)
        workbook = load_workbook(io.BytesIO(result))
        sheet = workbook["Axis Loads"]

        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        assert rows[0][5] is None
        assert rows[0][6] is None
        assert rows[0][7] is None

    def test_header_is_bold(self) -> None:
        robot = RobotInfo(
            model="KR 240 R2900",
            general=GeneralInfo(backup_name="Test"),
            controller=ControllerInfo(controller_type=ControllerType.UNKNOWN),
            warnings=WarningLog(),
        )
        result = ExcelExporter().export(robot)
        workbook = load_workbook(io.BytesIO(result))
        sheet = workbook["Axis Loads"]

        assert sheet["A1"].font.bold is True


class TestExportToFile:
    """Test the inherited export_to_file() with a real .xlsx file."""

    def test_export_to_file_produces_openable_workbook(
        self, sample_robot_info: RobotInfo, tmp_path: Path
    ) -> None:
        target = tmp_path / "output.xlsx"
        ExcelExporter().export_to_file(sample_robot_info, target)

        assert target.exists()
        workbook = load_workbook(target)
        assert "Summary" in workbook.sheetnames
