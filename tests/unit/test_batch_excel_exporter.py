"""Unit tests for BatchExcelExporter."""

import io
from pathlib import Path

from openpyxl import load_workbook

from kuka_value.exporters.batch_excel_exporter import BatchExcelExporter
from kuka_value.models.axis_load import AxisLoad
from kuka_value.models.batch_result import BatchItemResult
from kuka_value.models.controller_info import ControllerInfo, ControllerType
from kuka_value.models.general_info import GeneralInfo
from kuka_value.models.payload import Vector3D
from kuka_value.models.robot_info import RobotInfo
from kuka_value.models.warnings import WarningLog


class TestBatchExcelExport:
    def test_export_returns_bytes(self, sample_batch_results: list[BatchItemResult]) -> None:
        result = BatchExcelExporter().export(sample_batch_results)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_creates_three_sheets(self, sample_batch_results: list[BatchItemResult]) -> None:
        result = BatchExcelExporter().export(sample_batch_results)
        workbook = load_workbook(io.BytesIO(result))
        assert workbook.sheetnames == ["Summary", "Payloads", "Axis Loads"]

    def test_summary_has_one_row_per_backup(
        self, sample_batch_results: list[BatchItemResult]
    ) -> None:
        result = BatchExcelExporter().export(sample_batch_results)
        workbook = load_workbook(io.BytesIO(result))
        sheet = workbook["Summary"]

        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 3

    def test_summary_marks_success_as_ok(self, sample_batch_results: list[BatchItemResult]) -> None:
        result = BatchExcelExporter().export(sample_batch_results)
        workbook = load_workbook(io.BytesIO(result))
        sheet = workbook["Summary"]

        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        assert rows[0] == ("TestBackup", "KR 240 R2900", 2, 0, 1, "OK")

    def test_summary_marks_failure(self, sample_batch_results: list[BatchItemResult]) -> None:
        result = BatchExcelExporter().export(sample_batch_results)
        workbook = load_workbook(io.BytesIO(result))
        sheet = workbook["Summary"]

        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        failed_row = rows[2]
        assert failed_row[0] == "BrokenBackup"
        assert failed_row[1] == "-"
        assert "FAILED" in failed_row[5]

    def test_summary_header_is_bold(self, sample_batch_results: list[BatchItemResult]) -> None:
        result = BatchExcelExporter().export(sample_batch_results)
        workbook = load_workbook(io.BytesIO(result))
        assert workbook["Summary"]["A1"].font.bold is True

    def test_payloads_sheet_excludes_failed_backup(
        self, sample_batch_results: list[BatchItemResult]
    ) -> None:
        result = BatchExcelExporter().export(sample_batch_results)
        workbook = load_workbook(io.BytesIO(result))
        sheet = workbook["Payloads"]

        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 3
        backup_names = {r[0] for r in rows}
        assert backup_names == {"TestBackup", "SecondBackup"}

    def test_payload_row_prefixed_with_backup_and_model(
        self, sample_batch_results: list[BatchItemResult]
    ) -> None:
        result = BatchExcelExporter().export(sample_batch_results)
        workbook = load_workbook(io.BytesIO(result))
        sheet = workbook["Payloads"]

        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        second_backup_row = next(r for r in rows if r[0] == "SecondBackup")
        assert second_backup_row[1] == "KR 6 R900"
        assert second_backup_row[3] == 5.0

    def test_axis_loads_sheet_empty_when_none_present(
        self, sample_batch_results: list[BatchItemResult]
    ) -> None:
        result = BatchExcelExporter().export(sample_batch_results)
        workbook = load_workbook(io.BytesIO(result))
        sheet = workbook["Axis Loads"]

        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        assert rows == []

    def test_axis_loads_sheet_prefixed_with_backup_and_model(self) -> None:
        robot = RobotInfo(
            model="KR 240 R2900",
            general=GeneralInfo(backup_name="WithAxisLoad"),
            controller=ControllerInfo(controller_type=ControllerType.UNKNOWN),
            axis_loads=[
                AxisLoad(axis=3, mass=12.5, center_of_gravity=Vector3D(x=50.0, y=0.0, z=0.0))
            ],
            warnings=WarningLog(),
        )
        results = [BatchItemResult(source_path=Path("WithAxisLoad"), robot=robot, error=None)]

        result = BatchExcelExporter().export(results)
        workbook = load_workbook(io.BytesIO(result))
        sheet = workbook["Axis Loads"]

        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        assert rows[0][0] == "WithAxisLoad"
        assert rows[0][1] == "KR 240 R2900"
        assert rows[0][2] == 3  # axis number
        assert rows[0][3] == 12.5  # mass

    def test_empty_results_produces_header_only(self) -> None:
        result = BatchExcelExporter().export([])
        workbook = load_workbook(io.BytesIO(result))
        summary_rows = list(workbook["Summary"].iter_rows(min_row=2, values_only=True))
        assert summary_rows == []

    def test_export_to_file(
        self, sample_batch_results: list[BatchItemResult], tmp_path: Path
    ) -> None:
        target = tmp_path / "batch.xlsx"
        BatchExcelExporter().export_to_file(sample_batch_results, target)

        assert target.exists()
        workbook = load_workbook(target)
        assert "Summary" in workbook.sheetnames
